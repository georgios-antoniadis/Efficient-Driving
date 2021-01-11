import pygame
import time
import numpy as np
import neat
import os
from neatPlayer import player
from roadTerrain import road
import xlsxwriter
import os
import os.path
from os import path
import openpyxl

import matplotlib.pyplot as plt

carPng = pygame.image.load('carImage.png')
background = pygame.image.load('background.jpg')

pygame.init()
w, h = 1000,600 #Dimension of my window
win = pygame.display.set_mode((w,h)) #Touple as dimensions
pygame.display.set_caption('Efficient Driving Normal')
clock = pygame.time.Clock()

gen = 0

data = [[[]]]

#Pressure points (For acceleration and breaking use)
gasPressurePoints = 100
breakingPressurePoints = 100

#Font
statFont = pygame.font.SysFont('comicsans',50)

#Road
road = road()
goal = 500 #1km range which the cars have to travel
currentDir = os.getcwd()

def drawWindow(win,cars,gen):
    win.fill((0,0,255))
    for x in range(len(cars)):
        cars[x].draw(win,x)
    
    '''
    Drawing lines for checkpoints
    '''
    for i in range(50,550,50):
        pygame.draw.line(win,(255,0,0),[i,0],[i,600])

    #Displaying the current generation
    text = statFont.render('Gen: ' + str(gen), 1, (255,255,255))
    win.blit(text,(10,10))

for index in range(1,26):

    def main(genomes, config):
        global gen
        gen += 1
        cars = []
        nets = [] #Neural network
        ge = [] #Genome

        for _, g in genomes:
            net = neat.nn.FeedForwardNetwork.create(g, config)
            nets.append(net)
            #Initializing the cars
            cars.append(player(goal,road))
            g.fitness = 0
            ge.append(g)
        
        win = pygame.display.set_mode((w,h))
        clock = pygame.time.Clock()

        run = True

        #Making sure the car does not decide inhumanly fast
        reflexIndex = 1*60

        while run:

            events = pygame.event.get()
            for event in events:
                #Exiting and saving the data acquired
                if event.type == pygame.QUIT:
                    run = False
                    pygame.quit()
                    quit()

            driversRemaining = 0
            for c,car in enumerate(cars):    
                #Getting the output of our function
                """
                Inputs: 
                -Current velocity 
                -Current rpm
                -Current gear
                -Throttle or not and by what amount
                -Breaking or not and by what amount
                -Upper road speed limit
                -Lower road speed limit
                -Pressure applied on the throttle

                Output: 
                -The first value of the output function is fed into the accelerating or the breaking
                functions of the car object.
                    -The throttle/breaking is then analog to that value.
                    -A value less than 0 results in breaking
                    -A value more than 0 results in accelerating
                -The second value of the output function is controlling the gearing of the car
                    -A value more than 0.5 shifts the gear up
                    -A value of less than -0.5 shifts a gear down

                """            
                if car.getAlive():
                    if np.mod(reflexIndex,0.5*60) == 0:
                        #Controlling the pedals    
                        outputs = nets[c].activate((car.gasPressure,car.velocity,car.gearRatioIndex,car.rpm,road.speedlimitLower,road.speedlimitUpper))
                        #Throttle
                        #if outputs[0] > 0: #Accelerating
                        if outputs[0] > 1:
                            outputs[0] = outputs[0]/10
                        if outputs[0] > 10:
                            outputs[0] = outputs[0]/100
                        car.accelerating(outputs[0],gasPressurePoints)
                        #elif outputs[0] < 0: #Breaking
                            #car.accelerating(outputs[0],breakingPressurePoints)
                    if np.mod(reflexIndex,2*60) == 0:
                        outputs = nets[c].activate((car.gasPressure,car.velocity,car.gearRatioIndex,car.rpm,road.speedlimitLower,road.speedlimitUpper))
                        if outputs[1] > 0.5:   
                            car.shifting('up')
                        if outputs[1] < 0:
                            car.shifting('down')
                
            for c, car in enumerate(cars):
                if car.getAlive():
                    '''
                    #Road profile change
                    if car.position > 100:
                        car.road.incline = True
                        car.road.theta = 5
                        
                    if car.position > 150:
                        car.road.incline = False
                        car.road.theta = 0

                    if car.position > 170:
                        car.road.incline = True
                        car.road.theta = -5
                        
                    if car.position > 220:
                        car.road.incline = False
                        car.road.theta = 0    
                    '''

                    driversRemaining += 1
                    car.maxVel()
                    car.move() #Moving the car
                    car.fuelUse() #Fuel usage monitoring

                    #Fail safe for cases were the car does not move at all
                    if car.velocity == 0:
                        car.velFail += 1
                    elif car.velocity < road.speedlimitLower:
                        car.velFail += 1
                    elif car.velocity > road.speedlimitUpper:
                        car.velFail += 1
                    
                    if car.gearRatioIndex < 1:
                        car.alive = False
                    
                    if car.velFail == 4*60:
                        car.alive = False 
                    if car.velocity < 0:
                        car.alive = False
                    

                    #Fail safe for stalling
                    if car.rpm < car.idleRpm:
                        car.alive = False

                    #Checking that the car is within limits and monitoring its speed
                    if car.velocity < road.speedlimitLower:
                        pass
                    elif car.velocity <= 0: #Preventing it from not moving
                        pass
                    elif car.velocity > road.speedlimitUpper:
                        pass 
                    elif car.velocity >= road.speedlimitLower and car.velocity <= road.speedlimitUpper:
                        ge[c].fitness += 7
                    
                    if np.mod(car.position,50) == 0:
                        ge[c].fitness += (car.position/50 * 5)
                    
                    ge[c].fitness -= round(car.fuelUsed * (10**5)) #Reward for keeping fuel down

                    if gen > len(data):
                        data.append([])

                    if c+1 > len(data[gen-1]):
                        data[gen-1].append([])
                    
                    data[gen-1][c].append((round(car.velocity*1/car.velocityConversion),round(car.position),round(car.rpm*60*60),round(car.fuelUsed*10**5),car.gasPressure,car.gearRatioIndex,ge[c].fitness))

            if driversRemaining == 0:
                #If all drivers have finished their course start a new generation
                break

            #Updating the redflex index
            reflexIndex += 1

            drawWindow(win,cars,gen)

            pygame.display.update()
            clock.tick(60)

    def dataLog(data,currentDir):
        #Writing data in a csv file
        path = 'dataLog' + str(index) + '.csv'
        f = open(path,'w')

        f.write('ID, Distance, Average RPM, Total Fuel Used, Average Velocity, Average Throttle Pressure, Average Gear, Fitness')

        for i in range(len(data)):

            positions = []
            rpm = []
            fuel = []
            vel = []
            pressure = []
            gearIndex = []
            fitness = []

            f.write('\n')
            f.write('Generation ' + str(i))
            f.write('\n')
            f.write('\n')
            for j in range(len(data[i])):
                outstring = ''
                if j+1 > len(positions):
                    positions.append([])
                    rpm.append([])
                    fuel.append([])
                    vel.append([])
                    pressure.append([])
                    gearIndex.append([])
                    fitness.append([])

                for velocity, pos, revs, fuelUsed, gasP, gearId, fit in data[i][j]:
                    positions[j].append(pos)
                    rpm[j].append(revs)
                    fuel[j].append(fuelUsed)
                    vel[j].append(velocity)
                    pressure[j].append(gasP)
                    gearIndex[j].append(gearId)
                    fitness[j].append(fit)

                if float(np.max(fitness[j])) > 12000: #Plotting only the best performing individuals
                    #RPM plot
                    plt.figure()
                    plt.plot(positions[j],rpm[j],'b--',markersize=3)
                    plt.xlabel('Distance (m)')
                    plt.ylabel('RPM')
                    plt.savefig('Best performer rpm ' + str(np.max(fitness[j])) + '.png')
                    plt.close()
                    
                    #Throttle plot
                    plt.figure()
                    plt.plot(positions[j],pressure[j],'r--',markersize=3)
                    plt.xlabel('Distance (m)')
                    plt.ylabel('Throttle pressure (0-1)')
                    plt.savefig('Best performer throttle ' + str(np.max(fitness[j])) + '.png')
                    plt.close()

                    #Throttle plot
                    plt.figure()
                    plt.plot(positions[j],fuel[j],'g--',markersize=3)
                    plt.xlabel('Distance (m)')
                    plt.ylabel('Fuel used')
                    plt.savefig('Best performer fuel ' + str(np.max(fitness[j])) + '.png')
                    plt.close()

                    if os.path.exists(currentDir + 'bestperformers.csv') == False:
                        f1 = open('bestperformers.csv','w')
                    else:
                        f1 = open('bestperformers.csv','a')

                    f1.write(str(np.max(fitness[j])))
                    
                    f1.write('\n')
                    
                    for r in rpm[j]:
                        f1.write(str(r))
                        f1.write(',')
                    
                    f1.write('\n')
                    
                    for pr in pressure[j]:
                        f1.write(str(pr))
                        f1.write(',')
                    
                    f1.write('\n')
                    
                    for fu in fuel[j]:
                        f1.write(str(fu))
                        f1.write(',')

                    f1.write('\n')

                #Producing the string that will be outputted
                outstring += str(j) + ',' + str(round(np.max(positions[j]))) + ',' +str(round(np.average(rpm[j]))) + ',' + str(round(np.sum(fuel[j]))) + ',' + str(round(np.average(vel[j]))) + ',' + str(round(np.average(pressure[j]),3)) + ',' + str(round(np.average(gearIndex[j]),3)) + ',' + str(np.max(fitness[j]))
                f.write(outstring)
                f.write('\n')
        
        f.close()

    def saveData(index,data,currentDir):
        workbook = openpyxl.load_workbook(currentDir + r'\Data folder\\rawdata.xlsx')
        newSheet = workbook.create_sheet()
        newSheet.title = 'Run' + str(index)
        myRow = 1
        myCol = 1
        i = 0
        j = 0
        newSheet.cell(row = myRow,column = myCol).value = 'Gen'
        newSheet.cell(row = myRow,column = myCol+1).value = 'ID'
        newSheet.cell(row = myRow,column = myCol+2).value = 'Velocity'
        newSheet.cell(row = myRow,column = myCol+3).value = 'Position'
        newSheet.cell(row = myRow,column = myCol+4).value = 'RPM'
        newSheet.cell(row = myRow,column = myCol+5).value = 'Fuel Used'
        newSheet.cell(row = myRow,column = myCol+6).value = 'Throttle Pressure'
        newSheet.cell(row = myRow,column = myCol+7).value = 'Gear'
        myRow += 1
        for i in range(len(data)):
            for j in range(len(data[i])):
                for velocity, pos, revs, fuelUsed, p, gearID in data[i][j]:
                    newSheet.cell(row = myRow,column = myCol).value = i
                    newSheet.cell(row = myRow,column = myCol+1).value = j
                    newSheet.cell(row = myRow,column = myCol+2).value = velocity
                    newSheet.cell(row = myRow,column = myCol+3).value = pos
                    newSheet.cell(row = myRow,column = myCol+4).value = revs
                    newSheet.cell(row = myRow,column = myCol+5).value = fuelUsed
                    newSheet.cell(row = myRow,column = myCol+6).value = p
                    newSheet.cell(row = myRow,column = myCol+7).value = gearID
                    myRow += 1
        workbook.save(currentDir + r'\Data folder\\rawdata.xlsx')
        workbook.close()

    def myFun(x):
        return 0.01*(x) if x < 0 else np.max([0,x])

    def run(configFile):
        config = neat.config.Config(neat.DefaultGenome, neat.DefaultReproduction,
                            neat.DefaultSpeciesSet, neat.DefaultStagnation,
                            configFile)

        config.genome_config.add_activation('my_function', myFun)

        p = neat.Population(config)

        #Will show us statistics in console
        p.add_reporter(neat.StdOutReporter(True))
        stats = neat.StatisticsReporter()
        p.add_reporter(stats)

        winner = p.run(main,150) #Fintess function, generations
        print(winner)

    if __name__ == "__main__":
        localDir = os.path.dirname(__file__)
        configPath = os.path.join(localDir, 'config.txt')
        run(configPath)

    dataLog(data,currentDir)
    data = [[[]]]
    index += 1
    gen = 0
'''
#Script that saves all the necessary data
if path.exists(currentDir + r'\Data folder\\rawdata.xlsx') == False:
    workbook = xlsxwriter.Workbook(currentDir + r'\Data folder\\rawdata.xlsx')
    worksheet = workbook.add_worksheet()
    row = 0
    col = 0
    i = 0
    j = 0
    worksheet.write(row,col, 'Gen')
    worksheet.write(row,col+1, 'ID')
    worksheet.write(row,col+2, 'Velocity')
    worksheet.write(row,col+3, 'position')
    worksheet.write(row,col+4, 'RPM')
    worksheet.write(row,col+5, 'Fuel Used')
    row += 1
    for i in range(len(data)):
        for j in range(len(data[i])):
            for velocity, pos, revs, fuelUsed, pressure, gear in data[i][j]:
                worksheet.write(row,col, i)
                worksheet.write(row,col+1, j)
                worksheet.write(row,col+2, velocity)
                worksheet.write(row,col+3, pos)
                worksheet.write(row,col+4, revs)
                worksheet.write(row,col+5, fuelUsed)
                worksheet.write(row,col+6, pressure)
                worksheet.write(row,col+7, gear)
                row += 1
            workbook.close()
            
else:
    workbook = openpyxl.load_workbook(currentDir + r'\Data folder\\rawdata.xlsx')
    names = workbook.sheetnames
    index = len(names) + 1
    saveData(index,data,currentDir)
'''
pygame.quit()
        
